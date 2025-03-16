#!/usr/bin/env python
# coding: utf-8

# Reading the image file and test xlsx file

# In[12]:


import cv2, pandas as pd, matplotlib.pyplot as plt

img = cv2.cvtColor(cv2.imread("C:\\Users\\HP\\OneDrive\\Desktop\\Counting shape near text\\input.jpg"), cv2.COLOR_BGR2RGB)
plt.imshow(img)
plt.axis('off')
plt.show()

display(pd.read_excel("C:\\Users\\HP\\OneDrive\\Desktop\\Counting shape near text\\test.xlsx").head())


# 

# 

# Fetching the images from the excel sheet to Symbol folder

# In[13]:


import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
from io import BytesIO


excel_path = 'C:\\Users\\HP\\OneDrive\\Desktop\\Counting shape near text\\test.xlsx'
symbols_folder = 'C:\\Users\\HP\\OneDrive\\Desktop\\Counting shape near text\\symbol'

# Need to create a folder to store the symbols 
os.makedirs(symbols_folder, exist_ok=True)

# Load the tst.xlxs
wb = load_workbook(excel_path)
ws = wb.active


image_mapping = {}  

for img in ws._images:
    if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
        anchor = img.anchor._from 
        row = anchor.row + 1      
        col = anchor.col + 1
        if hasattr(img, '_data'):
            img_bytes = img._data()
        else:
            img_bytes = img.ref  
        image_mapping[(row, col)] = img_bytes

for row in range(2, ws.max_row + 1):
    key = (row, 2)  
    if key in image_mapping:
        try:
            img_data = image_mapping[key]
            image = Image.open(BytesIO(img_data))
            image.save(os.path.join(symbols_folder, f"symbol_row{row}.png"))
            print(f"Symbol image saved: symbol_row{row}.png")
        except Exception as e:
            print(f" Failed to save symbol at row {row}: {e}")
    else:
        print(f" No symbol image found in row {row}")


# In[ ]:





# Creating Augmented data for each symbols (50-60 with different rotations and angles)

# In[15]:


import numpy as np
from tensorflow.keras.preprocessing.image import ImageDataGenerator

symbols_folder = "C:\\Users\\HP\\OneDrive\\Desktop\\Counting shape near text\\symbols"
augmented_folder = "C:\\Users\\HP\\OneDrive\\Desktop\\Counting shape near text\\augmented_data"
os.makedirs(augmented_folder, exist_ok=True)

datagen = ImageDataGenerator(
    rotation_range=20,
    zoom_range=0.2,
    width_shift_range=0.1,
    height_shift_range=0.1,
    brightness_range=(0.6, 1.4),
    fill_mode='nearest'
)

for symbol_img in os.listdir(symbols_folder):
    label = os.path.splitext(symbol_img)[0].replace(" ", "_")
    label_folder = os.path.join(augmented_folder, label)
    os.makedirs(label_folder, exist_ok=True)

    img = cv2.imread(os.path.join(symbols_folder, symbol_img))
    img = cv2.resize(img, (64, 64))
    img = np.expand_dims(img, axis=0)

    i = 0
    for batch in datagen.flow(img, batch_size=1, save_to_dir=label_folder,
                              save_prefix=label, save_format='png'):
        i += 1
        if i > 50:  
            break


# In[ ]:





# In[ ]:





# In[ ]:


Creating the masks of images


# In[ ]:


augmented_folder = "C:\\Users\\HP\\OneDrive\\Desktop\\Counting shape near text\\augmented_data"
mask_folder = "C:\\Users\\HP\\OneDrive\\Desktop\\Counting shape near text\\masks"
os.makedirs(mask_folder, exist_ok=True)

# Iterate through each label folder
for label_folder in os.listdir(augmented_folder):
    label_path = os.path.join(augmented_folder, label_folder)
    mask_label_folder = os.path.join(mask_folder, label_folder)
    os.makedirs(mask_label_folder, exist_ok=True)

    for img_name in os.listdir(label_path):
        img_path = os.path.join(label_path, img_name)
        img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)

        # Thresholding to create a binary mask
        _, mask = cv2.threshold(img, 128, 255, cv2.THRESH_BINARY)

        # Save mask
        mask_path = os.path.join(mask_label_folder, img_name)
        cv2.imwrite(mask_path, mask)

print("Masks created successfully!")


# In[ ]:





# In[ ]:





# In[ ]:


Creating the model


# In[21]:


from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Conv2D, MaxPooling2D, Flatten, Dense, Dropout, BatchNormalization, Input
from tensorflow.keras.preprocessing.image import ImageDataGenerator
import os

symbol_folder = "C:\\Users\\HP\\OneDrive\\Desktop\\Counting shape near text\\augmented_data"

num_classes = len([d for d in os.listdir(symbol_folder) if os.path.isdir(os.path.join(symbol_folder, d))])

model = Sequential([
    Input(shape=(64, 64, 3)),
    Conv2D(64, (3, 3), activation='relu', padding='same'),
    BatchNormalization(),
    MaxPooling2D((2, 2)),

    Conv2D(128, (3, 3), activation='relu', padding='same'),
    BatchNormalization(),
    MaxPooling2D((2, 2)),

    Conv2D(256, (3, 3), activation='relu', padding='same'),
    BatchNormalization(),
    MaxPooling2D((2, 2)),

    Conv2D(512, (3, 3), activation='relu', padding='same'),
    BatchNormalization(),
    MaxPooling2D((2, 2)),

    Flatten(),
    Dense(512, activation='relu'),
    Dropout(0.5),
    Dense(num_classes, activation='softmax')
])

model.compile(optimizer='adam', loss='categorical_crossentropy', metrics=['accuracy'])

datagen = ImageDataGenerator(rescale=1./255, validation_split=0.2)

train_data = datagen.flow_from_directory(
    symbol_folder,
    target_size=(64, 64),
    batch_size=32,
    class_mode='categorical',
    subset='training'
)

val_data = datagen.flow_from_directory(
    symbol_folder,
    target_size=(64, 64),
    batch_size=32,
    class_mode='categorical',
    subset='validation'
)

model.fit(train_data, validation_data=val_data, epochs=35)

model.save("C:\\Users\\HP\\OneDrive\\Desktop\\Counting shape near text\\better_symbol_classifier_v4.h5")


# In[ ]:





# In[24]:


from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing.image import img_to_array
from tqdm import tqdm
import pytesseract

# Load model
model = load_model("symbol_classifier_cnn.h5")

# Rebuild label map
label_map = {v: k for k, v in model.class_indices.items()} if hasattr(model, 'class_indices') else {
    0: "Damper1", 1: "Damper2", 2: "Linear_Supply_Diffuser", 3: "Linear_Return_Diffuser",
    4: "Supply_Diffuser", 5: "Return_Diffuser", 6: "Thermostat"
}

# Load image
image_path = "C:\\Users\\HP\\OneDrive\\Desktop\\Counting shape near text\\test1.jpg"
image = cv2.imread(image_path)
gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

# Extract text using pytesseract
text_data = pytesseract.image_to_data(gray, output_type=pytesseract.Output.DICT)

# Store text box coordinates
text_boxes = []
for i in range(len(text_data['text'])):
    if text_data['text'][i].strip():  # Only include non-empty text
        x, y, w, h = text_data['left'][i], text_data['top'][i], text_data['width'][i], text_data['height'][i]
        text_boxes.append((x, y, w, h, text_data['text'][i]))

# Symbol detection logic
H, W = gray.shape
step = 48
win_size = 64
empty_thresh = 30

counts = {label: 0 for label in label_map.values()}
output_image = image.copy()
patches = []
coords = []
nearest_texts = []  # New list to store nearest text info

for y in range(0, H - win_size, step):
    for x in range(0, W - win_size, step):
        roi = gray[y:y+win_size, x:x+win_size]
        if np.mean(roi) > empty_thresh:
            roi_color = image[y:y+win_size, x:x+win_size]
            roi_resized = cv2.resize(roi_color, (64, 64))
            roi_array = img_to_array(roi_resized) / 255.0
            patches.append(roi_array)
            coords.append((x, y))

print(f"Total patches to classify: {len(patches)}")
patches = np.array(patches)

# Classification and nearest text detection
BATCH_SIZE = 64
for i in tqdm(range(0, len(patches), BATCH_SIZE), desc="Classifying patches"):
    batch = patches[i:i+BATCH_SIZE]
    batch_coords = coords[i:i+BATCH_SIZE]
    preds = model.predict(batch)

    for pred, (x, y) in zip(preds, batch_coords):
        class_idx = np.argmax(pred)
        confidence = pred[class_idx]

        if confidence > 0.85:
            label = label_map[class_idx]
            counts[label] += 1

            # Find nearest text
            nearest_text = "N/A"
            min_distance = float('inf')
            for (tx, ty, tw, th, text) in text_boxes:
                distance = np.sqrt((x - tx)**2 + (y - ty)**2)
                if distance < min_distance:
                    min_distance = distance
                    nearest_text = text

            nearest_texts.append(nearest_text)

            cv2.rectangle(output_image, (x, y), (x+win_size, y+win_size), (0, 255, 0), 1)
            cv2.putText(output_image, label, (x, y-5), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 255, 0), 1)

cv2.imwrite("C:\\Users\\HP\\OneDrive\\Desktop\\Counting shape near text\\cad1.png", output_image)

# Save results
df = pd.DataFrame({
    "Component Label": list(counts.keys()),
    "Count": list(counts.values()),
    "Nearest Text": nearest_texts
})

df.to_excel("C:\\Users\\HP\\OneDrive\\Desktop\\Counting shape near text\\final1.xlsx", index=False)

print("Symbol count report with nearest text detection saved successfully.")


# In[ ]:





# In[ ]:


Finding threholding and diffrent iamges with counting 


# In[19]:


import matplotlib.pyplot as plt
from tensorflow.keras.models import load_model
import tensorflow as tf
from tensorflow.keras.preprocessing.image import img_to_array
from collections import Counter

def detect_symbols_in_cad(cad_image_path, model_path, symbol_folder, output_dir=None):
    
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    print("Loading model...")
    model = load_model(model_path)
    
    class_names = sorted([d for d in os.listdir(symbol_folder) 
                         if os.path.isdir(os.path.join(symbol_folder, d))])
    print(f"Found {len(class_names)} symbol classes: {class_names}")
    
    print(f"Loading CAD image: {cad_image_path}")
    img = cv2.imread(cad_image_path)
    if img is None:
        raise ValueError(f"Failed to load image at {cad_image_path}")
    
    img_color = img.copy()
    
    if len(img.shape) == 3:
        img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    else:
        img_gray = img
    
    if output_dir:
        cv2.imwrite(os.path.join(output_dir, "01_input_gray.png"), img_gray)
    
    print("Preprocessing image...")
    _, img_thresh = cv2.threshold(img_gray, 200, 255, cv2.THRESH_BINARY_INV)
    
    kernel = np.ones((3, 3), np.uint8)
    img_clean = cv2.morphologyEx(img_thresh, cv2.MORPH_OPEN, kernel, iterations=1)
    
    if output_dir:
        cv2.imwrite(os.path.join(output_dir, "02_threshold.png"), img_thresh)
        cv2.imwrite(os.path.join(output_dir, "03_cleaned.png"), img_clean)
    
    print("Finding potential symbol regions...")
    contours, _ = cv2.findContours(img_clean, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    print(f"Found {len(contours)} potential regions")
    
    min_contour_area = 20  
    contours = [c for c in contours if cv2.contourArea(c) > min_contour_area]
    print(f"After filtering: {len(contours)} regions")
    
    if output_dir:
        contour_img = img_color.copy()
        cv2.drawContours(contour_img, contours, -1, (0, 255, 0), 2)
        cv2.imwrite(os.path.join(output_dir, "04_contours.png"), contour_img)
    
    def predict_symbol(image_patch):
        resized = cv2.resize(image_patch, (64, 64))
        
        if len(resized.shape) == 2:
            resized = cv2.cvtColor(resized, cv2.COLOR_GRAY2RGB)
        elif resized.shape[2] == 4: 
            resized = cv2.cvtColor(resized, cv2.COLOR_RGBA2RGB)
        
        normalized = resized.astype('float32') / 255.0
        
        input_tensor = np.expand_dims(normalized, axis=0)
        
        prediction = model.predict(input_tensor, verbose=0)[0]
        class_idx = np.argmax(prediction)
        confidence = prediction[class_idx]
        
        return class_names[class_idx], confidence
    
    print("Classifying symbols...")
    detections = []
    padding = 5 
    
    patches_dir = os.path.join(output_dir, "symbol_patches") if output_dir else None
    if patches_dir and not os.path.exists(patches_dir):
        os.makedirs(patches_dir)
    
    for i, contour in enumerate(contours):
        x, y, w, h = cv2.boundingRect(contour)
        
        x1 = max(0, x - padding)
        y1 = max(0, y - padding)
        x2 = min(img_gray.shape[1], x + w + padding)
        y2 = min(img_gray.shape[0], y + h + padding)
        
        symbol_patch = img_gray[y1:y2, x1:x2]
        
        if symbol_patch.size == 0:
            continue
        
        if patches_dir:
            cv2.imwrite(os.path.join(patches_dir, f"patch_{i:04d}.png"), symbol_patch)
        
        symbol_name, confidence = predict_symbol(symbol_patch)
        
        min_confidence = 0.5
        if confidence >= min_confidence:
            detections.append({
                'symbol': symbol_name,
                'confidence': confidence,
                'box': (x1, y1, x2, y2)
            })
    
    print(f"Detected {len(detections)} symbols")
    
    def merge_overlapping_boxes(detections, iou_threshold=0.3):
        """Merge overlapping boxes of the same symbol class"""
        if not detections:
            return []
        
        sorted_dets = sorted(detections, key=lambda x: x['confidence'], reverse=True)
        merged_dets = []
        
        while sorted_dets:
            best = sorted_dets.pop(0)
            merged_dets.append(best)
            
            to_keep = []
            for det in sorted_dets:
                if det['symbol'] != best['symbol']:
                    to_keep.append(det)
                    continue
                
                box1 = best['box']
                box2 = det['box']
                
                x1 = max(box1[0], box2[0])
                y1 = max(box1[1], box2[1])
                x2 = min(box1[2], box2[2])
                y2 = min(box1[3], box2[3])
                
                if x2 <= x1 or y2 <= y1:
                    to_keep.append(det)
                    continue
                
                inter_area = (x2 - x1) * (y2 - y1)
                
                box1_area = (box1[2] - box1[0]) * (box1[3] - box1[1])
                box2_area = (box2[2] - box2[0]) * (box2[3] - box2[1])
                union_area = box1_area + box2_area - inter_area
                
                iou = inter_area / union_area
                
                if iou < iou_threshold:
                    to_keep.append(det)
            
            sorted_dets = to_keep
        
        return merged_dets
    
    merged_detections = merge_overlapping_boxes(detections)
    print(f"After merging: {len(merged_detections)} unique symbols")
    
    if output_dir:
        result_img = img_color.copy()
        
        np.random.seed(42) 
        colors = {}
        for name in class_names:
            colors[name] = tuple(map(int, np.random.randint(0, 255, 3)))
        
        for det in merged_detections:
            x1, y1, x2, y2 = det['box']
            symbol = det['symbol']
            conf = det['confidence']
            color = colors.get(symbol, (0, 255, 0))
            
            cv2.rectangle(result_img, (x1, y1), (x2, y2), color, 2)
            
            label = f"{symbol}: {conf:.2f}"
            text_size = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.5, 2)[0]
            
            cv2.rectangle(result_img, 
                          (x1, y1 - 20), 
                          (x1 + text_size[0], y1),
                          color, -1)
            
            cv2.putText(result_img, label, (x1, y1 - 5),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 2)
        
        cv2.imwrite(os.path.join(output_dir, "result_detection.png"), result_img)
    
    symbol_counts = Counter([det['symbol'] for det in merged_detections])
    
 
    
    return merged_detections, symbol_counts

def train_focused_model(symbol_folder, cad_images_folder, output_model_path):
   
    pass

def main():
    model_path = "C:\\Users\\HP\\OneDrive\\Desktop\\Counting shape near text\\better_symbol_classifier_v4.h5"
    symbol_folder = "C:\\Users\\HP\\OneDrive\\Desktop\\Counting shape near text\\augmented_data"
    cad_image_path = "C:\\Users\\HP\\OneDrive\\Desktop\\Counting shape near text\\input.jpg"
    output_dir = "C:\\Users\\HP\\OneDrive\\Desktop\\Counting shape near text\\cad_detection_results"

    
    detections, counts = detect_symbols_in_cad(
        cad_image_path=cad_image_path,
        model_path=model_path,
        symbol_folder=symbol_folder,
        output_dir=output_dir
    )
    
   

if __name__ == "__main__":
    main()


# In[ ]:




